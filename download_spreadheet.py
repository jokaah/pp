#!/usr/bin/env python3
import csv
import io
import re
import sys
import time
import pathlib
from urllib.parse import urlparse

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from openpyxl import load_workbook


ONLY_TABS = [
    "Top Series",
    "Top Games",
    "#A",
    "B",
    "C",
    "D",
    "EF",
    "G",
    "HI",
    "JK",
    "L",
    "M",
    "NO",
    "PQ",
    "R",
    "S",
    "So",
    "T",
    "UVXYZ",
    "W"
]


def extract_spreadsheet_id(s: str) -> str:
    if "/" not in s and len(s) >= 20:
        return s
    parsed = urlparse(s)
    parts = [p for p in parsed.path.split("/") if p]
    if "spreadsheets" in parts and "d" in parts:
        return parts[parts.index("d") + 1]
    raise ValueError("Could not extract spreadsheet ID. Paste URL or the raw ID.")


def safe_filename(name: str) -> str:
    name = name.strip().replace("/", "-").replace("\\", "-").replace(":", " -")
    name = re.sub(r'[<>:"|?*]', "_", name)
    return (name[:200] or "Sheet")


def make_session() -> requests.Session:
    s = requests.Session()
    s.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/126.0 Safari/537.36"
    })

    retry = Retry(
        total=8,
        connect=8,
        read=8,
        backoff_factor=1.0,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=("GET", "HEAD"),
        raise_on_status=False,
    )
    adapter = HTTPAdapter(max_retries=retry, pool_connections=10, pool_maxsize=10)
    s.mount("https://", adapter)
    s.mount("http://", adapter)
    return s


def download_xlsx_to_file(session: requests.Session, spreadsheet_id: str, dest_path: pathlib.Path) -> None:
    """
    Stream the XLSX to disk with long read timeout and progress reporting.
    The read timeout is per socket read; as long as data trickles in, it won't time out.
    """
    url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"

    # (connect_timeout, read_timeout) — tune these as needed.
    # For huge spreadsheets, a big read timeout is normal.
    timeout = (30, 600)  # 10 minutes per-chunk

    dest_path.parent.mkdir(parents=True, exist_ok=True)

    print(f"[i] Requesting XLSX export…")
    with session.get(url, allow_redirects=True, stream=True, timeout=timeout) as r:
        r.raise_for_status()

        ct = (r.headers.get("content-type") or "").lower()
        if "text/html" in ct:
            # Often indicates it's not actually public / interstitial / consent page
            raise RuntimeError(
                "Got HTML instead of XLSX. The sheet may not be public, "
                "or Google returned an interstitial page."
            )

        total_len = r.headers.get("content-length")
        total_len = int(total_len) if total_len and total_len.isdigit() else None

        tmp_path = dest_path.with_suffix(dest_path.suffix + ".part")

        downloaded = 0
        t0 = time.time()
        last_print = 0.0

        with open(tmp_path, "wb") as f:
            for chunk in r.iter_content(chunk_size=1024 * 1024):  # 1MB
                if not chunk:
                    continue
                f.write(chunk)
                downloaded += len(chunk)

                now = time.time()
                if now - last_print >= 1.0:
                    last_print = now
                    mb = downloaded / (1024 * 1024)
                    speed = mb / max(now - t0, 0.001)

                    if total_len:
                        pct = (downloaded / total_len) * 100
                        tot_mb = total_len / (1024 * 1024)
                        print(f"[i] {mb:.1f}/{tot_mb:.1f} MB ({pct:.1f}%) @ {speed:.1f} MB/s")
                    else:
                        print(f"[i] {mb:.1f} MB downloaded @ {speed:.1f} MB/s")

        tmp_path.replace(dest_path)
        print(f"[✓] Saved XLSX: {dest_path}")


def xlsx_to_csvs(
    xlsx_path: pathlib.Path,
    out_dir: pathlib.Path,
    include_hidden: bool = True,
    only_tabs: list[str] | None = None,
) -> None:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    out_dir.mkdir(parents=True, exist_ok=True)

    # Normalize allowlist (case-insensitive, trimmed)
    allow = None
    if only_tabs:
        allow = {t.strip().casefold() for t in only_tabs if t.strip()}

    used = set()
    found = set()

    for ws in wb.worksheets:
        if not include_hidden and ws.sheet_state != "visible":
            continue

        if allow is not None:
            title_key = ws.title.strip().casefold()
            if title_key not in allow:
                continue
            found.add(title_key)

        filename = safe_filename(ws.title)
        base = filename
        i = 2
        while filename.lower() in used:
            filename = f"{base} ({i})"
            i += 1
        used.add(filename.lower())

        out_path = out_dir / f"{filename}.csv"
        print(f"[i] Writing {out_path.name}")

        with out_path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(["" if v is None else v for v in row])

    if allow is not None:
        missing = sorted(allow - found)
        if missing:
            print("[!] These tabs were requested but not found:")
            for name in missing:
                print(f"    - {name}")

    print("[✓] CSV export complete.")


def main():
    if len(sys.argv) < 2:
        print("Usage: python download_spreadsheet_xlsx.py <SPREADSHEET_URL_OR_ID> [out_dir]")
        sys.exit(1)

    spreadsheet_input = sys.argv[1]
    out_dir = pathlib.Path(sys.argv[2]) if len(sys.argv) >= 3 else pathlib.Path("./csv_exports")
    out_dir.mkdir(parents=True, exist_ok=True)

    ssid = extract_spreadsheet_id(spreadsheet_input)
    print(f"[i] Spreadsheet ID: {ssid}")

    session = make_session()

    xlsx_path = out_dir / f"{ssid}.xlsx"
    if xlsx_path.exists():
        print(f"[i] XLSX already exists, reusing: {xlsx_path}")
    else:
        download_xlsx_to_file(session, ssid, xlsx_path)

    xlsx_to_csvs(xlsx_path, out_dir, include_hidden=True, only_tabs=ONLY_TABS)
    print("[✓] Done.")


if __name__ == "__main__":
    main()
